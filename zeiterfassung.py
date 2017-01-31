# encoding=utf8
import sys
import openpyxl
import collections

reload(sys)
sys.setdefaultencoding('utf8')

# TODO Wymienic CHANGE-ME na sciezke w ktorej znajduje sie katalog z excelami pracownikow. Ostatni slash jest potrzebny
path_to_workhours_folder = 'C:/CHANGEME/'

wb = openpyxl.load_workbook(filename='employee-list.xlsx')
ws = wb.get_sheet_by_name('Lista pracownikow')


def format_name(name):
    """ Konwertuje z 'Imie Nazwisko' na 'imie.nazwisko' """
    return name.replace(' ', '.').lower()


def format_path(name):
    """ Generuje sciezke do pliku z excelem pracownika """
    return path_to_workhours_folder + name + '/' + name + '-zeiterfassung.xlsx'


def extract_project_names(worksheet):
    project_names = []
    for row_idx in range(3, worksheet.max_row + 1):
        project_name = worksheet.cell(row=row_idx, column=3).value
        if project_name is not None:
            project_names.append(project_name)
    return project_names


# TODO na razie to wypisuje wyniki na ekran. Zrozumiec jak dziala openpyxl i zapisac je do excela :)
def process_employee(name):
    """ Przetwarza godziny pracownika """
    print name
    name = format_name(name)
    employee_wb = openpyxl.load_workbook(filename=format_path(name), data_only=True)
    projects_dict = {}
    # dla kazdego arkusza w pliku
    is_first_ws = True
    for worksheet in employee_wb:
        # jesli pierwszy, to wyciagamy liste projektow
        if is_first_ws:
            project_names = extract_project_names(worksheet)
            # dla kazdego projektu, ustawiamy jego ilosc godzin na 0
            for project_name in project_names:
                add_project_hours(projects_dict, project_name, 0)
            # koniec obslugi pierwszego, przetwarzamy pozostale
            is_first_ws = False
        else:
            for row_idx in range(2, worksheet.max_row + 1):
                project_name = worksheet.cell(row=row_idx, column=1).value
                hours = worksheet.cell(row=row_idx, column=worksheet.max_column).value
                # jesli znalazl projekt i godziny nad nim spedzone, dodajemy
                if project_name is not None and hours is not None:
                    add_project_hours(projects_dict, project_name, hours)
    # sortujemy slownik
    workhours = collections.OrderedDict(sorted(projects_dict.items()))
    for item in workhours.items():
        print '\t', item[0] + ': ' + str(item[1])


def add_project_hours(dict, name, hours):
    if name not in dict:
        dict[name] = hours
    else:
        dict[name] = dict[name] + hours


def main():
    # zbieramy liste pracownikow
    for row_idx in range(2, ws.max_column - 1):
        name = ws.cell(row=row_idx, column=1).value
        process_employee(name)


if __name__ == "__main__":
    main()
